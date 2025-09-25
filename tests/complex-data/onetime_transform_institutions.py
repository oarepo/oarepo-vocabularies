#
# Copyright (c) 2025 CESNET z.s.p.o.
#
# This file is a part of oarepo-vocabularies (see https://github.com/oarepo/oarepo-vocabularies).
#
# oarepo-vocabularies is free software; you can redistribute it and/or modify it
# under the terms of the MIT License; see LICENSE file for more details.
#
from __future__ import annotations

import re

import openpyxl
from unidecode import unidecode

# nasty
rowidx = 0


def next_row(it):
    global rowidx  # noqa: PLW0603
    rowidx += 1
    return [x.value for x in next(it)]


def empty(r):
    return all(not val for val in r)


wb_obj = openpyxl.load_workbook("tests/complex-data/institutions_old.xlsx")
sheet_obj = wb_obj.active


it = sheet_obj.iter_rows()

stack = []
ids = set()

shortcuts = {
    "fakulta-jaderna-a-fyzikalne-inzenyrska": "FJFI",
    "prirodovedecka": "Pr",
    "pedagogicka": "Ped",
    "filozoficka": "F",
    "strojni": "St",
}

try:
    row = next_row(it)
    while empty(row):
        row = next_row(it)
    while not empty(row):
        row = next_row(it)
    while empty(row):
        row = next_row(it)
    header = True
    while True:
        if not empty(row):
            if header:
                assert row[0] == "level"
                assert row[1] == "slug"
                assert row[7] == "props.acronym"

                sheet_obj.cell(row=rowidx, column=1).value = "hierarchy.parent"
                sheet_obj.cell(row=rowidx, column=2).value = "id"
                header = False
                row = next_row(it)
                continue

            level = int(row[0])
            slug = re.sub(r"\W+", "-", row[2].lower().strip())
            while level <= len(stack):
                stack.pop()
            if stack:
                parent, base = stack[-1]
            else:
                parent, base = ("", "")
            id_ = str(row[7] or slug)
            split_id = []
            if id_.endswith("-cas"):
                split_id = [id_]
            else:
                for word in id_.split("-"):
                    if word in shortcuts:
                        split_id.append(shortcuts[word])
                        continue
                    for ci, c in enumerate(word):
                        if ci == 0 or c == c.upper():
                            split_id.append(c.upper())
            short_id = "".join(split_id)

            if base:
                id_ = f"{base}-{id_}"
                short_id = f"{base}-{short_id}"

            base = short_id if short_id not in ids and len(short_id) > 2 else id_
            id_ = unidecode(id_.lower())
            assert id_ not in ids

            ids.add(id_)
            stack.append((id_, base))
            sheet_obj.cell(row=rowidx, column=1).value = parent
            sheet_obj.cell(row=rowidx, column=2).value = id_

        row = next_row(it)
except StopIteration:
    pass


wb_obj.save("tests/complex-data/institutions.xlsx")
